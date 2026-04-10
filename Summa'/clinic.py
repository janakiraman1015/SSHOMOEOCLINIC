from flask import Flask,render_template,redirect,request,flash,url_for,jsonify,session
import os
from openpyxl import Workbook,load_workbook
from datetime import datetime
import calendar
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
 

app = Flask(__name__)
app.secret_key = "@@123@@567"
CURR_DIR = os.getcwd()
DATA_DIR = os.path.join(CURR_DIR,"data")

# username and password
sender_email = "mailtosshomoeoclinic@gmail.com"
password = "mbqegiruvftuqmhj"

app.config["SESSION_PERMANENT"] = False     # Sessions expire when the browser is closed
app.config["SESSION_TYPE"] = "filesystem"


def send_mail_smtp(receiver_mail,body,subject):
    receiver_email = receiver_mail
    # Create the email
    message = MIMEMultipart("alternative")
    message["Subject"] = subject
    message["From"] = sender_email
    message["To"] = receiver_email

    # Email body
    text = body
    part = MIMEText(body, "html")
    message.attach(part)

    # Connect to Gmail SMTP server
    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()  # Secure connection
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, message.as_string())

    print("Email sent successfully!")


def excel_file_handler(excel_file,data):
    if os.path.exists(excel_file):
        wbook = load_workbook(excel_file)
        sheet = wbook.active
        last_row = sheet.max_row

        for row in sheet.iter_rows(min_row=2,values_only=True):
            if(row[3] == data["phone_number"] and row[4]== data["date"]):
                flash("Apponitment already booked!!","error")
                return redirect(url_for("home"))

        last_num = int(sheet.cell(row=last_row,column=1).value)
        sheet.append([str(last_num+1),data["patient_name"],data["email"],data["phone_number"],data["date"],data["problem"],"Pending"])

    else:
        wbook = Workbook()
        sheet = wbook.active
        headers = ["SI No","Patient Name","Email","Phone Number","Date","Problem","Status"]
        sheet.append(headers)
        sheet.append(["1",data["patient_name"],data["email"],data["phone_number"],data["date"],data["problem"],"Pending"])
    
    wbook.save(excel_file)


def get_overall_patients():
    if  os.path.exists(DATA_DIR):
        files = [f for f in os.listdir(DATA_DIR)
                    if os.path.isfile(os.path.join(DATA_DIR, f))]
        patients_count = 0
        for file in files:
            file_path = os.path.join(DATA_DIR,file)
            wb_data = load_workbook(file_path)
            ws_data = wb_data.active
            for row in ws_data.iter_rows(min_row=2,values_only=True):
                if(row[6] == "Accepted"):
                    patients_count+=1

        return patients_count
        
    
    else:
        return 0
    

@app.route("/api/get/barchart_data")
def barchart_data():
     files = [f for f in os.listdir(DATA_DIR)
                    if os.path.isfile(os.path.join(DATA_DIR, f))]
     
     curr_year = datetime.now().year
     months = []
     appointments = []
     patients  = []
     for file in files:
         filename = file.split(".")[0]
         if(str(filename.split("-")[1]) == str(curr_year)):
             splitted_month = int(filename.split("-")[0])
             months.append(calendar.month_abbr[splitted_month])
             print("file:",file)
             file_path = os.path.join(DATA_DIR,file)
             wb_data = load_workbook(file_path)
             ws_data = wb_data.active
             appointments.append(ws_data.max_row-1)
             count = 0
             for row in ws_data.iter_rows(values_only=True,min_row=2):
                 if(row[6] == "Accepted"):
                     count+=1
             patients.append(count)

     print("patients:",patients)

     data = {
            "labels":months,
            "appointments":appointments,
            "patients": patients
        }
     return jsonify(data)

# # logoin
# from functools import wraps

# def login_required(f):
#     @wraps(f)
#     def wrapper(*args, **kwargs):
#         if 'user' not in session:
#             return redirect(url_for('login'))
#         return f(*args, **kwargs)
#     return wrapper



@app.route("/",methods=['GET','POST'])
@app.route("/home",methods=['GET','POST'])
def home():
    if request.method == "POST":
        data = request.form.to_dict()
        print("data:",data)
        os.makedirs(DATA_DIR,exist_ok=True)

        date = data['date']
        app_month = date.split('-')[1]
        app_year = date.split("-")[0]
        print(app_month,app_year)
        excel_file = os.path.join(DATA_DIR,f"{app_month}-{app_year}.xlsx")
        excel_file_handler(excel_file,data)
        return render_template("home.html")
    else:
        session["name"] = None
        return render_template("home.html")

@app.route("/admin",methods=['GET','POST'])
def admin():
    if(request.method == 'POST'):
        data = request.form.to_dict()
        if(data['user'] == 'admin' and data['password'] == 'clinic@123@'):
            session["name"] = "admin"
            return redirect(url_for('dashboard'))
        else:
            flash("Invalid credentials","error")
            return redirect(url_for('admin'))

    else:
        session["name"] = None
        print("session cleared!!")
        return render_template("Admin.html")


def calculate_mom_for_month(month, year):
    try:
        folder_path = DATA_DIR

        if month is None or year is None:
            return {"error": "Please provide both month and year."}

        current_key = f"{month:02d}-{year}"

        if month == 1:
            prev_month = 12
            prev_year = year - 1
        else:
            prev_month = month - 1
            prev_year = year

        previous_key = f"{prev_month:02d}-{prev_year}"

        current_file = None
        previous_file = None

        # -------- Find files --------
        for file in os.listdir(folder_path):
            if file.endswith(".xlsx"):
                if file.startswith(current_key):
                    current_file = file
                elif file.startswith(previous_key):
                    previous_file = file

        if not current_file:
            return {"error": f"Current month file not found: {current_key}.xlsx"}

        # -------- Helper --------
        def get_score(file):
            file_path = os.path.join(folder_path, file)

            wb = load_workbook(file_path)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                return {"accepted": 0, "rejected": 0, "total": 0, "score": 0}

            headers = [str(h).strip().lower() for h in rows[0]]

            if "status" not in headers:
                return {"error": "No 'status' column found"}

            status_idx = headers.index("status")

            total = 0
            accepted = 0
            rejected = 0

            for row in rows[1:]:
                if not any(row):
                    continue

                total += 1
                status = str(row[status_idx]).strip().lower() if row[status_idx] else ""

                if status in ["accepted"]:
                    accepted += 1
                elif status in ["declined"]:
                    rejected += 1

            conversion_rate = ((total - rejected) / total) *100
            return {
                "accepted": accepted,
                "rejected": rejected,
                "total": total,
                "conversion_rate": conversion_rate
            }

        # -------- Current --------
        current_stats = get_score(current_file)
        if "error" in current_stats:
            return current_stats

        current_acc = current_stats["accepted"]
        current_total = current_stats["total"]
        current_conversion = current_stats["conversion_rate"]

        # -------- Previous --------
        if previous_file:
            previous_stats = get_score(previous_file)
            if "error" in previous_stats:
                return previous_stats

            previous_acc = previous_stats["accepted"]
            previous_total = previous_stats["total"] 
            previous_conversion = previous_stats["conversion_rate"]   

            accept_change = ((current_acc - previous_acc)/previous_acc )*100
            total_change = ((current_total - previous_total)/previous_total )*100
            conversion_change = ((current_conversion - previous_conversion)/previous_conversion )*100
            
        else:
            accept_change = None
            total_change = None

        # -------- Final Output --------
        return {
            "current_file": current_file,
            "previous_file": previous_file,

            "current_accepted": current_stats["accepted"],
            "current_rejected": current_stats["rejected"],
            "current_total": current_stats["total"],

            "previous_accepted": previous_stats["accepted"] if previous_stats else None,
            "previous_rejected": previous_stats["rejected"] if previous_stats else None,
            "previous_total": previous_stats["total"] if previous_stats else None,

            "patients_percent": round(accept_change, 2) if accept_change is not None else None,
            "appointments_percent": round(total_change, 2) if total_change is not None else None,
            "conversion_rates_change_percent": round(conversion_change, 2) if conversion_change is not None else None
        }

    except Exception as e:
        return {"error": str(e)}
    

@app.route('/dashboard')
def dashboard():
    if not session.get("name"):
        print("session:",session)
        if os.path.exists(DATA_DIR):

            selected_month = request.args.get('month')
            selected_year = request.args.get('year')

            now = datetime.now()

            # Default to current month/year
            if not selected_month:
                selected_month = now.strftime("%m")
            if not selected_year:
                selected_year = str(now.year)

            files = [f for f in os.listdir(DATA_DIR)
                    if os.path.isfile(os.path.join(DATA_DIR, f))]

            data = []
            patients_data = []

            # Load all files for tables
            for file in files:
                file_path = os.path.join(DATA_DIR, file)
                wb_data = load_workbook(file_path)
                ws_data = wb_data.active

                for row in ws_data.iter_rows(min_row=2, values_only=True):
                    if row[6] == "Pending":
                        data.append(list(row))
                    elif row[6] == "Accepted":
                        patients_data.append(list(row))

            # 🔥 Selected month file
            current_file = os.path.join(DATA_DIR, f"{selected_month}-{selected_year}.xlsx")

            total_appointments = 0
            patients = 0

            if os.path.exists(current_file):
                wb = load_workbook(current_file)
                ws = wb.active

                total_appointments = ws.max_row - 1

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[6] == "Accepted":
                        patients += 1

            # Avoid crash
            conversion_rate = int((patients / total_appointments) * 100) if total_appointments > 0 else 0

            overall_patients = get_overall_patients()
            deviation = calculate_mom_for_month(int(selected_month), int(selected_year))

            return render_template(
                "Dashboard.html",
                pending_data=data,
                patients_data=patients_data,
                total_appointments=total_appointments,
                patients=patients,
                conversion_rate=conversion_rate,
                overall_patients=overall_patients,
                deviation=deviation,
                selected_month=selected_month,
                selected_year=selected_year
            )

        return render_template("Dashboard.html")
    else:
        return redirect(url_for("admin"))

def send_decline_appoint_mail(patient_name, id, date, patient_mail):
    try:
        subject = "“Appointment Decline from SS Homoeo Clinic"

        body_template = """
       <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">

                <p>Dear <b>{PatientName}</b>,</p>

                <p>Thank you for choosing <b>{ClinicName}</b>.</p>

                <p>
                We regret to inform you that your appointment request for 
                <b>{AppointmentDate}</b> could not be confirmed at this time due to scheduling constraints.
                </p>

                <p style="color: #d9534f;"><b>
                We sincerely apologize for any inconvenience this may cause.
                </b></p>

                <h3>🔄 What You Can Do Next:</h3>
                <ul>
                <li>You may request a different date that suits your availability.</li>
                <li>Alternatively, contact us directly and we will help you find the next available slot.</li>
                </ul>

                <h3>📞 Need Assistance?</h3>
                <p>
                Phone: {ClinicPhone}<br>
                Email: {ClinicEmail}
                </p>

                <p>
                We value your time and trust, and we would be happy to assist you with a new appointment.
                </p>

                <p>
                Warm regards,<br>
                <b>{ClinicName} Team</b>
                </p>

                <hr>

                <p style="font-size: 12px; color: gray;">
                This is an automated message. Please do not reply directly to this email.
                </p>

            </body>
</html>
        """

        body = body_template.format(
            PatientName=patient_name,
            ClinicName="SS Homoeo Clinic",
            AppointmentDate=date,
            ClinicPhone="+91 9443011830",
            ClinicEmail="mailtosshomoeoclinic@gmail.com",
            ClinicTagline="Your Health, Our Priority"
        )

        send_mail_smtp(patient_mail, body, subject)

    except Exception as e:
        print("Error occurred when sending decline appointment mail:", e)



@app.route('/decline_appointment',methods=['POST'])
def decline_appointment():
    data = request.get_json()
    declined_row = data['row']
    print("row:",declined_row)
    date = declined_row[4]
    app_month = date.split('-')[1]
    app_year = date.split("-")[0]
    excel_file = os.path.join(DATA_DIR,f"{app_month}-{app_year}.xlsx")
    if os.path.exists(excel_file):
        wb_data = load_workbook(excel_file)
        ws_data = wb_data.active
        for row in ws_data.iter_rows(min_row=2):
            if(row[3].value == declined_row[3] and row[4].value == declined_row[4]):
                row[6].value = "Declined"
        wb_data.save(excel_file)
        send_decline_appoint_mail(declined_row[1],declined_row[0],declined_row[4],declined_row[2])
    else:
        print('Excel file not exists!!')
    return jsonify({"success":"true"})



def send_accept_appoint_mail(patient_name, id, date, patient_mail):
    try:
        subject = "Appointment Confirmation from SS Homoeo Clinic"

        body_template = """
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6;">
                <p>Dear <b>{PatientName}</b>,</p>

                <p>Thank you for choosing <b>{ClinicName}</b>. Your appointment has been successfully scheduled.</p>

                <h3>🗓 Appointment Details:</h3>
                <ul>
                    <li><b>Patient Name:</b> {PatientName}</li>
                    <li><b>Appointment ID:</b> {AppointmentID}</li>
                    <li><b>Date:</b> {AppointmentDate}</li>
                    <li><b>Doctor:</b> {DoctorName}</li>
                    <li><b>Clinic Location:</b> {ClinicAddress}</li>
                </ul>

                <h3>📌 Important Instructions:</h3>
                <ul>
                    <li>Bring any previous medical records or prescriptions (if applicable).</li>
                    <li>If you need to reschedule or cancel, inform us at least 24 hours in advance.</li>
                </ul>

                <h3>📞 Need Help?</h3>
                <p>
                    Phone: {ClinicPhone}<br>
                    Email: {ClinicEmail}
                </p>

                <p>We look forward to providing you with the best care.</p>

                <p>
                    Warm regards,<br>
                    <b>{ClinicName} Team</b><br>
                    <i>{ClinicTagline}</i>
                </p>

                <hr>
                <p style="font-size: 12px; color: gray;">
                    This is an automated confirmation email. Please do not reply directly to this message.
                </p>
            </body>
        </html>
        """

        body = body_template.format(
            PatientName=patient_name,
            ClinicName="SS Homoeo Clinic",
            AppointmentID=id,
            AppointmentDate=date,
            DoctorName="Dr. S. Saraswathi , Dr. S. Srimathi",
            ClinicAddress="Coma stores, Kangeyan Kovil street, 43/2, S Bazaar Rd, Kovilpatti, Tamil Nadu 628501",
            ClinicPhone="+91 9443011830",
            ClinicEmail="mailtosshomoeoclinic@gmail.com",
            ClinicTagline="Your Health, Our Priority"
        )

        send_mail_smtp(patient_mail, body, subject)

    except Exception as e:
        print("Error occurred when sending appointment mail:", e)

@app.route('/accept_appointment',methods=['POST'])
def accept_appointment():
    data = request.get_json()
    declined_row = data['row']
    print("declined_row:",declined_row)
    date = declined_row[4]
    app_month = date.split('-')[1]
    app_year = date.split("-")[0]
    excel_file = os.path.join(DATA_DIR,f"{app_month}-{app_year}.xlsx")
    if os.path.exists(excel_file):
        wb_data = load_workbook(excel_file)
        ws_data = wb_data.active
        for row in ws_data.iter_rows(min_row=2):
            if(row[3].value == declined_row[3] and row[4].value == declined_row[4]):
                row[6].value = "Accepted"
        wb_data.save(excel_file)

        send_accept_appoint_mail(declined_row[1],declined_row[0],declined_row[4],declined_row[2])
    else:
        print('Excel file not exists!!')
    return jsonify({"success":"true"})

if __name__ == "__main__":
    app.run(debug=True)